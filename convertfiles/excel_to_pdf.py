import streamlit as st
import openpyxl
from openpyxl import load_workbook  # Ensure this import statement is correct
import xlrd
from fpdf import FPDF
import tempfile
import os
import datetime
import socket
import pyodbc


# Function to connect to the database
def connect_to_db():
    conn = pyodbc.connect(
        "Driver={ODBC Driver 18 for SQL Server};"
        "Server=Z3PHYR\\SQLEXPRESS;"
        "Database=PowergridDB;"
        "UID=sa;"
        "PWD=powergrid;"
        "Encrypt=no;"
    )
    return conn

st.logo("Images/powergrid.webp")


# Function to insert file metadata into the database
def insert_file_data(name, description, status, content_type, size, updated_by, updated_on, updated_from):
    try:
        conn = connect_to_db()
        if conn:
            cursor = conn.cursor()

            # Convert Python types to SQL Server expected types
            updated_on_sql = updated_on.strftime('%Y-%m-%d %H:%M:%S.%f')[:-3] if isinstance(updated_on, datetime.datetime) else updated_on
            size = int(size)  # Convert size to integer
            status = int(status)  # Convert status to integer

            # Call the stored procedure with parameters in correct order
            cursor.execute("{CALL sp_InsertFile (?, ?, ?, ?, ?, ?, ?, ?)}",
                           (name, description, status, content_type, size, updated_by, updated_on_sql, updated_from))

            conn.commit()

            # Display success message (optional)
            #st.success("File metadata stored successfully.")
          
    except pyodbc.Error as e:
        st.error(f"Error inserting file data: {e}")
        # More detailed error logging
        st.write("SQL Error:", e)
        st.write("SQL State:", e.args[0])
        st.write("SQL Message:", e.args[1])
    except Exception as ex:
        st.error(f"Unexpected error: {ex}")
    finally:
        if conn:
            conn.close()

# Function to convert Excel to PDF with specified columns per page
def convert_excel_to_pdf(excel_file, columns_per_page=5):
    _, ext = os.path.splitext(excel_file.name)
    
    if ext.lower() == ".xls":
        return convert_xls_to_pdf(excel_file, columns_per_page)
    elif ext.lower() == ".xlsx":
        return convert_xlsx_to_pdf(excel_file, columns_per_page)
    else:
        raise ValueError("Unsupported file format. Please upload an Excel file in .xls or .xlsx format.")

# Function to check if a file is a valid XLS (Excel 97-2003) file
def is_xls(file):
    # XLS magic number (header)
    xls_magic_number = b'\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1'

    # Read the first 8 bytes of the file
    first_bytes = file.read(8)
    file.seek(0)  # Reset file pointer to beginning

    # Check if the file header matches the XLS magic number
    return first_bytes.startswith(xls_magic_number)

# Function to check if a file is a valid XLSX (Excel 2007 or later) file
def is_xlsx(file):
    # XLSX magic number (header)
    xlsx_magic_number = b'\x50\x4B\x03\x04'

    # Read the first 4 bytes of the file
    first_bytes = file.read(4)
    file.seek(0)  # Reset file pointer to beginning

    # Check if the file header matches the XLSX magic number
    return first_bytes == xlsx_magic_number

def convert_xls_to_pdf(excel_file, columns_per_page):
    workbook = xlrd.open_workbook(file_contents=excel_file.read())
    pdf = FPDF()

    for sheet_name in workbook.sheet_names():
        worksheet = workbook.sheet_by_name(sheet_name)
        num_cols = worksheet.ncols
        num_rows = worksheet.nrows

        # Calculate number of pages needed based on columns per page
        num_pages = (num_cols + columns_per_page - 1) // columns_per_page

        for page_num in range(num_pages):
            pdf.add_page()

            # Set font for PDF
            pdf.set_font("Arial", size=12)

            # Custom page layout (adjust margins as needed)
            pdf.set_left_margin(10)
            pdf.set_right_margin(10)

            # Determine columns to include on this page
            start_col = page_num * columns_per_page
            end_col = min(start_col + columns_per_page, num_cols)

            # Iterate through rows and columns for the selected range
            for row in range(num_rows):
                for col in range(start_col, end_col):
                    cell_value = worksheet.cell_value(row, col)
                    if cell_value is not None:
                        pdf.cell(38, 11, str(cell_value), border=1)
                pdf.ln()  # Move to the next line after each row

    # Save PDF to a temporary file
    pdf_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf").name
    pdf.output(pdf_file)

    return pdf_file

def convert_xlsx_to_pdf(excel_file, columns_per_page):
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    pdf = FPDF()

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        num_cols = worksheet.max_column
        num_rows = worksheet.max_row

        # Calculate number of pages needed based on columns per page
        num_pages = (num_cols + columns_per_page - 1) // columns_per_page

        for page_num in range(num_pages):
            pdf.add_page()

            # Set font for PDF
            pdf.set_font("Arial", size=12)

            # Custom page layout (adjust margins as needed)
            pdf.set_left_margin(10)
            pdf.set_right_margin(10)

            # Determine columns to include on this page
            start_col = page_num * columns_per_page + 1
            end_col = min(start_col + columns_per_page - 1, num_cols)

            # Handle the first column separately (for serial numbers with no header)
            if start_col == 1:
                pdf.cell(38, 11, "", border=1)  # Empty cell for the first column

            # Iterate through rows and columns for the selected range
            for row in range(1, num_rows + 1):
                for col in range(start_col, end_col + 1):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value is not None:
                        pdf.cell(38, 11, str(cell_value), border=1)
                pdf.ln()  # Move to the next line after each row

    # Save PDF to a temporary file
    pdf_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf").name
    pdf.output(pdf_file)


    return pdf_file


# Streamlit app
def app():

    st.set_page_config(
    page_title="निर्मल: XLS to PDF",
    page_icon="🔄",
    layout="centered",
    initial_sidebar_state="expanded"
    )

    st.markdown(
        """
        <style>
            [data-testid=stSidebar] [data-testid=stImage]{
                text-align: center;
                display: block;
                margin-left: auto;
                margin-right: auto;
                width: 100%;
            }
        </style>
        """, unsafe_allow_html=True
    )

    with st.sidebar:
        st.image('./Images/logo-white.png', width=150)

    with st.sidebar:
        st.markdown(f"""<h6 style='text-align: center;'>Powergrid &copy; {datetime.datetime.now().year} All rights reserved.</h6>""", unsafe_allow_html=True)

    if not st.session_state.authenticated:
        st.error(" 🔒 Please log in before accessing the resource.")
        return
    st.markdown("<h1 style='text-align: center;'><  निर्मल: XLS/X To PDF Converter  ></h1>", unsafe_allow_html=True)
    st.image('./Images/convertExcelToPDF.png', width=670)
    st.write("")
    # File uploader for Excel files
    uploaded_file = st.file_uploader("*Upload an Excel file that you want to convert.", type=["xls", "xlsx"], key="excel_uploader")

    if uploaded_file is not None:
        try:
            if is_xls(uploaded_file):
                st.success("XLS file Uploaded Successfully!")
                file_size = len(uploaded_file.getvalue())
    
            elif is_xlsx(uploaded_file):
                st.success("XLSX file Uploaded Successfully!")
                file_size = len(uploaded_file.getvalue())
            else:
                st.warning("Please upload a valid Excel file (XLS or XLSX).")
                return
             
            # Convert Excel to PDF with 5 columns per page
            pdf_file = convert_excel_to_pdf(uploaded_file, columns_per_page=5)

            # Provide download link for the converted PDF
            with open(pdf_file, "rb") as f:
                st.download_button(
                    label="Download PDF",
                    data=f,
                    file_name="converted.pdf",
                    mime="application/pdf"
                )

        except Exception as e:
            st.error(f"Error converting Excel to PDF: {e}")

        finally:
            # Clean up temporary files
            if 'pdf_file' in locals() and os.path.exists(pdf_file):
                os.remove(pdf_file)

        insert_file_data(
                    name=uploaded_file.name,
                    description="Excel to PDF",
                    status=1,  # Example status
                    content_type=os.path.splitext(uploaded_file.name)[1].lower(),  # Get file extension
                    size=file_size,
                    updated_by=st.session_state.username,  # Get current user's login name
                    updated_on=datetime.datetime.now(),  # Current date and time
                    updated_from=socket.gethostbyname(socket.gethostname())  # Example of retrieving hostname
                    )

if __name__ == "__page__":
    app()
